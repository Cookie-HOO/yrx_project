import os
import shutil
import typing

import pandas as pd
from openpyxl.reader.excel import load_workbook

from yrx_project.scene.split_table.const import SCENE_TEMP_PATH
from yrx_project.utils.file import copy_file

TEMP_FILE_PATH = os.path.join(SCENE_TEMP_PATH, "split_table.xlsx")


def sheets2excels():
    """
    将指定path的excel的多个sheet转成多个excel，保留所有格式
    TEMP_FILE_PATH 指向的excel的每个sheet会被保存为单独的excel文件
    新文件以sheet名命名，最后删除原始文件
    """
    try:
        # 加载原始工作簿（保留格式）
        wb = load_workbook(TEMP_FILE_PATH)
        # 获取目录路径
        dir_path = os.path.dirname(TEMP_FILE_PATH)
        # 为每个sheet创建单独的工作簿
        for sheet_name in wb.sheetnames:
            # 创建新工作簿
            new_wb = load_workbook(TEMP_FILE_PATH)
            # 删除新工作簿中不需要的sheet
            for sheet in new_wb.sheetnames:
                if sheet != sheet_name:
                    del new_wb[sheet]
            # 处理文件名中的非法字符
            safe_sheet_name = "".join(
                c for c in sheet_name if c.isalnum() or c in (' ', '_', '-')
            ).strip()
            # 设置输出路径
            output_path = os.path.join(dir_path, f"{safe_sheet_name}.xlsx")
            # 保存新工作簿
            new_wb.save(output_path)
        # 关闭原始工作簿
        wb.close()
        # 删除原始文件
        os.remove(TEMP_FILE_PATH)
        return True, f"成功分割为 {len(wb.sheetnames)} 个文件，保留原格式"
    except Exception as e:
        return False, f"错误: {str(e)}"


class SplitTable:
    def __init__(self, path: str, sheet_name_or_index: typing.Union[str, int], row_num_for_column: int,
                 raw_df: pd.DataFrame, reorder_dict: dict):
        """
        reorder_dict:
            reorder
                {"reorder_序号": True}  # 说明序号列需要被重置为1,2,3...
        """
        self.path = path
        self.sheet_name_or_index = sheet_name_or_index
        self.row_num_for_column = int(row_num_for_column)
        self.raw_df = raw_df
        self.excel_obj = None
        self.reorder_dict = reorder_dict  # 当有「序号」列时，用户选择的是否需要重新排序

        self.CUR_SHEET = sheet_name_or_index
        self.col_widths = None

    def init_env(self):
        # 1. 初始化路径, 拷贝临时文件
        if os.path.exists(SCENE_TEMP_PATH):
            shutil.rmtree(SCENE_TEMP_PATH)
        os.makedirs(SCENE_TEMP_PATH)
        temp_path = TEMP_FILE_PATH
        copy_file(self.path, temp_path)

        # 2. 生成 ExcelStyleValue, 删除其他sheet
        from yrx_project.utils.excel_style import ExcelStyleValue
        self.excel_obj = ExcelStyleValue(temp_path, self.sheet_name_or_index, run_mute=True)

        all_sheets = self.excel_obj.get_sheets_name()
        sheets_to_delete = [s for s in all_sheets if s != self.CUR_SHEET]
        self.excel_obj.batch_delete_sheet(sheets_to_delete)

        # 3. 删除当前sheet列名以下的行
        self.col_widths = self.excel_obj.get_cols_width()
        self.excel_obj.batch_delete_row(self.row_num_for_column + 1, 1_000_000)  # 删除从指定行号+1开始的所有行

    def copy_rows_to(self, name, group_value):
        """
        将符合条件的数据从原始 DataFrame 中复制到新的 sheet 中。

        :param name: 新的 sheet 名称
        :param group_value: 用于筛选数据的条件字典，例如 {"col1": "1", "col2": 2}
        """
        if self.excel_obj is None:
            raise ValueError("Run init_env first")

        # 1. 创建新 sheet（如果不存在）
        self.excel_obj.batch_copy_sheet([name], append=True, del_old=False)
        self.excel_obj.switch_sheet(name)

        # 2. 筛选符合条件的行
        condition = pd.Series(True, index=self.raw_df.index)  # 初始化条件为全 True
        for col, value in group_value.items():
            condition &= (self.raw_df[col] == value)  # 逐步叠加筛选条件

        filtered_df = self.raw_df[condition]  # 筛选出符合条件的行
        for column in filtered_df.columns:
            if self.reorder_dict.get("reorder_" + column):
                filtered_df[column] = range(1, len(filtered_df) + 1)  # 重新排序
        # 3. 将筛选出的行写入新 sheet
        start_row = self.row_num_for_column + 1  # 从指定行号的下一行开始写入
        for i, (_, row) in enumerate(filtered_df.iterrows()):
            values = row.tolist()  # 获取行数据
            self.excel_obj.set_row(start_row + i, values)  # 写入到 Excel 中

        self.excel_obj.set_cols_width(self.col_widths)  # 设置列宽
        # 4. 切换回主 sheet
        self.excel_obj.switch_sheet(self.CUR_SHEET)  # 复位

    def wrap_up(self):
        # 删除临时 sheet
        self.excel_obj.batch_delete_sheet([self.CUR_SHEET])

        # 保存并关闭 Excel 文件
        self.excel_obj.save()
