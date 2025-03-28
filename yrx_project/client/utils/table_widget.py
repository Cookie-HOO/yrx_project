import time
import typing

import pandas as pd
import xlsxwriter
from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtGui import QBrush, QColor
from PyQt5.QtWidgets import QTableWidgetItem, QWidget, QComboBox, QPushButton, QHBoxLayout, QTableWidget, QCheckBox, \
    QLineEdit, QSizePolicy, QLabel, QVBoxLayout, QHeaderView, QSpinBox
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook

from yrx_project.client.base import WindowWithMainWorkerBarely, BaseWorker
from yrx_project.client.const import COLOR_BLUE
from yrx_project.client.utils.button_color_widget import ColorPickerToolButton
from yrx_project.client.utils.cascader_selector import CascaderSelectComboBox
from yrx_project.client.utils.multi_selector import MultiSelectComboBox
from yrx_project.client.utils.number_widget import CustomSpinBox
from yrx_project.utils.color_util import rgb_to_hex
from yrx_project.utils.df_util import read_excel_file_with_multiprocessing, MergedCells

"""
pyqt5的table组件，的set和get 
"""

# 单元格填充的回调方法，返回QColor对象，QColor(255, 255, 255)
CELL_STYLE_FUNC_TYPE = typing.Callable[[pd.DataFrame, int, int], QColor]
# 单元格组件的回调方法，返回QComboBox对象或None，None表示普通文本
CELL_WIDGET_FUNC_TYPE = typing.Callable[[pd.DataFrame, int, int], typing.Union[None, QComboBox]]


def drag_enter_event(event):
    if event.mimeData().hasUrls():
        event.accept()
    else:
        event.ignore()


def drag_move_event(event):
    if event.mimeData().hasUrls():
        event.setDropAction(Qt.CopyAction)
        event.accept()
    else:
        event.ignore()


def drag_drop_event(event, func: typing.Callable[[typing.List[str]], typing.Any]):
    file_names = [url.toLocalFile() for url in event.mimeData().urls()]
    func(file_names)


class TableWidgetWrapper:
    def __init__(self, table_widget=None, del_rows_button=False, add_rows_button=False, drag_func=None):
        self.table_widget = table_widget or QTableWidget()
        # 隐藏指定列
        self.__hidden_column()
        # 添加按钮
        new_table = self.__add_buttons(add_rows_button, del_rows_button)
        if new_table is not None:
            self.table_widget = new_table
        # 支持将文件拖拽到这个table中
        if drag_func is not None:
            self.table_widget.setAcceptDrops(True)
            self.table_widget.dragEnterEvent = drag_enter_event
            self.table_widget.dropEvent = lambda event: drag_drop_event(event, drag_func)
            self.table_widget.dragMoveEvent = drag_move_event
        self.raw_df = None  # fill之后存储数据，可以用 self.raw_df.merged_cells获取合并单元格信息
        self.merged_cells = None  # [[min_row, min_col, max_row, max_col], [], []]
        self.has_column_widget = False  # 默认在header中是column，如果设置了column_func说明占了一行当column

    def __hidden_column(self):
        for i in range(self.table_widget.columnCount()):
            header = self.table_widget.horizontalHeaderItem(i)
            if header is not None:
                if header.text().startswith("__"):
                    self.table_widget.setColumnHidden(i, True)

    def __add_buttons(self, add_rows_button, del_rows_button):
        if not add_rows_button and not del_rows_button:
            return

        # 获取table所在的布局
        layout = self.table_widget.parent().layout()

        # 获取table在布局中的位置
        index = layout.indexOf(self.table_widget)

        # 创建一个水平布局，并添加两个按钮
        hbox = QHBoxLayout()

        # 创建两个按钮
        if add_rows_button:
            add_rows_button = QPushButton('末尾添加一项', self.table_widget.parent())
            add_rows_button.clicked.connect(self.__add_row)
            hbox.addWidget(add_rows_button)
        if del_rows_button:
            del_rows_button = QPushButton('删除选中项', self.table_widget.parent())
            del_rows_button.clicked.connect(self.__delete_row)
            hbox.addWidget(del_rows_button)

        # 创建一个新的table，并复制原来table的所有内容
        new_table = QTableWidget(self.table_widget.rowCount(), self.table_widget.columnCount(), self.table_widget.parent())
        for i in range(self.table_widget.rowCount()):
            for j in range(self.table_widget.columnCount()):
                new_table.setItem(i, j, self.table_widget.item(i, j))

        # 删除原来的table
        layout.removeWidget(self.table_widget)
        self.table_widget.setParent(None)

        # 在原来table的位置添加新的组合
        layout.insertLayout(index, hbox)
        layout.insertWidget(index + 1, new_table)
        return new_table

    def __add_row(self):
        row_count = self.table_widget.rowCount()
        # 在表格末尾添加一行
        self.table_widget.insertRow(row_count)

    def __delete_row(self):
        # 获取选中的行
        selected_rows = self.table_widget.selectionModel().selectedRows()
        # 从后往前删除，防止行数变化影响删除
        for index in sorted(selected_rows, reverse=True):
            self.table_widget.removeRow(index.row())

    def set_col_width(self, col_index: int, width: int):
        self.table_widget.setColumnWidth(col_index, width)
        return self

    def get_cell_value(self, row: int, column: int) -> typing.Optional[str]:
        widget = self.table_widget.item(row, column) or self.table_widget.cellWidget(row, column)
        if isinstance(widget, QTableWidgetItem): return widget.text()
        elif isinstance(widget, QComboBox): return widget.currentText()  # 下拉框
        elif isinstance(widget, QCheckBox): return widget.isChecked()  # 全局单选框
        elif isinstance(widget, (MultiSelectComboBox, CascaderSelectComboBox)): return widget.selected_values()  # 多选的值
        elif isinstance(widget, QSpinBox): return widget.value()  # 数字
        elif isinstance(widget, ColorPickerToolButton): return widget.color.name()  # 颜色：返回十六进制颜色字符串
        elif isinstance(widget, QWidget):
            line_edit = widget.findChild(QLineEdit)
            if line_edit is not None:
                return line_edit.text()
            label = widget.findChild(QLabel)
            if label is not None:
                return label.text()
        return None

    def get_cell_value_with_color(self, row: int, column: int) -> (typing.Optional[str], typing.Optional[str]):
        # 尝试获取QTableWidgetItem（普通文本）
        item = self.table_widget.item(row, column)
        if item is not None:
            background_color = item.background().color()
            if background_color.isValid():
                return item.text(), rgb_to_hex((background_color.red(), background_color.green(), background_color.blue()))
            else:
                return item.text(), None

        # 尝试获取QWidget（下拉框）
        widget = self.table_widget.cellWidget(row, column)
        if isinstance(widget, QComboBox):
            return widget.currentText(), None

        elif isinstance(widget, QCheckBox):  # 全局单选框
            return widget.isChecked(), None

        # 如果既不是QTableWidgetItem也不是QComboBox，返回None
        return None, None

    def fill_data_with_color(
            self,
            df: pd.DataFrame,
            cell_style_func: typing.Callable[[pd.DataFrame, int, int], QColor] = None,
            cell_widget_func: typing.Callable[[pd.DataFrame, int, int], QWidget] = None,
            column_widget_func: typing.Callable[['TableWidgetWrapper', list, int], QWidget] = None,  # 入参：df的列名 和 列索引
    ):
        cols_to_drop = [i for i in df.columns if str(i).startswith('__')]
        fill_df = df.drop(cols_to_drop, axis=1)
        self.raw_df = df
        if hasattr(df, "merged_cells"):
            self.merged_cells = df.merged_cells
        fill_df.fillna('', inplace=True)
        fill_data = fill_df.values.tolist()
        self.table_widget.setColumnCount(len(fill_data[0]) if fill_data else 0)
        start_value_row = 0  # 数据从第几行开始

        header_obj = self.table_widget.horizontalHeader()
        header_obj.setSectionResizeMode(QHeaderView.Interactive)  # 可自由调整列宽

        if column_widget_func is None:  # 正常设置header
            self.table_widget.setRowCount(len(fill_data))
            self.table_widget.setHorizontalHeaderLabels([str(i) for i in fill_df.columns])
        else:  # 用第一行当header
            self.has_column_widget = True
            self.table_widget.setRowCount(len(fill_data)+1)
            self.table_widget.setHorizontalHeaderLabels([str(ind+1) for ind, v in enumerate(fill_df.columns)])
            start_value_row = 1
            for index, col_name in enumerate(fill_df.columns):
                self.table_widget.setCellWidget(0, index, column_widget_func(self, list(fill_df.columns), index))

        self.table_widget.setUpdatesEnabled(False)
        # 设置数据
        for i, row in enumerate(fill_data):
            i = i + start_value_row
            for j, value in enumerate(row):
                item = QTableWidgetItem(str(value))
                if cell_widget_func is None:
                    if cell_style_func:
                        color = cell_style_func(fill_df, i, j)
                        if color:
                            item.setBackground(QBrush(color))
                    self.table_widget.setItem(i, j, item)
                else:
                    item = cell_widget_func(fill_df, i, j) or item
                    self.table_widget.setCellWidget(i, j, item)

        # 设置合并情况
        if self.merged_cells:
            for merged_cell in self.merged_cells.iter():
                min_row, min_col, max_row, max_col = merged_cell
                args = (min_row, min_col, max_row-min_row+1, max_col-min_col+1)
                self.table_widget.setSpan(*args)
        self.table_widget.setUpdatesEnabled(True)

        return self

    def get_data_as_df(self) -> pd.DataFrame:
        headers = []
        if not self.has_column_widget:
            for i in range(self.table_widget.columnCount()):
                header = self.table_widget.horizontalHeaderItem(i)
                if header is not None:
                    headers.append(header.text())
                else:
                    headers.append(f'Column{i}')
        else:  # 说明第一行才是
            headers = [self.get_cell_value(0, i) for i in range(self.table_widget.columnCount())]
        data = []
        data_start_row = 1 if self.has_column_widget else 0
        for i in range(data_start_row, self.table_widget.rowCount()):
            row_data = []
            for j in range(self.table_widget.columnCount()):
                item = self.get_cell_value(i, j)
                row_data.append(item or '')
            data.append(row_data)
        df = pd.DataFrame(data, columns=headers)
        if self.merged_cells:
            df.merged_cells = self.merged_cells
        return df

    def get_data_as_rows_and_color(self) -> (pd.DataFrame, list):
        headers = []
        for i in range(self.table_widget.columnCount()):
            header = self.table_widget.horizontalHeaderItem(i)
            if header is not None:
                headers.append(header.text())
            else:
                headers.append(f'Column{i}')
        data = []
        colors = []
        for i in range(self.table_widget.rowCount()):
            row_data = []
            color_data = []
            for j in range(self.table_widget.columnCount()):
                item, color = self.get_cell_value_with_color(i, j)
                row_data.append(item or '')
                color_data.append(color)
            data.append(row_data)
            colors.append(color_data)
        df = pd.DataFrame(data, columns=headers)
        return df, colors

    def clear(self):
        """全部清空，只保留表头"""
        self.table_widget.setRowCount(0)

    def clear_content(self):
        """清空内容，保留行列"""
        for i in range(self.table_widget.rowCount()):
            for j in range(self.table_widget.columnCount()):
                self.table_widget.setItem(i, j, QTableWidgetItem(""))

    # TODO: 上下交换的方法，在有widget的时候会报错，SIGSEGV 错误，暂时不好解决，可以通过同时记录所有表格内容，交换时彻底生成新内容解决
    def swap_rows(self, row1, row2):
        """交换指定的两行（包含所有列的内容和控件）"""
        # 边界检查
        if (
                row1 < 0
                or row1 >= self.table_widget.rowCount()
                or row2 < 0
                or row2 >= self.table_widget.rowCount()
        ):
            return

        # 遍历所有列（修正列遍历范围）
        for col in range(self.table_widget.columnCount()):  # 不交换最后一列
            # 交换 QTableWidgetItem
            item1 = self.table_widget.takeItem(row1, col)
            item2 = self.table_widget.takeItem(row2, col)
            self.table_widget.setItem(row1, col, item2)
            self.table_widget.setItem(row2, col, item1)

            # 交换 QWidget（安全方式）
            # 1. 获取控件引用
            widget1 = self.table_widget.cellWidget(row1, col)
            widget2 = self.table_widget.cellWidget(row2, col)

            # 2. 立即解除表格关联
            self.table_widget.setCellWidget(row1, col, None)
            self.table_widget.setCellWidget(row2, col, None)

            # 3. 操作控件（此时控件仅由Python引用保持存活）
            if widget1:
                widget1.setParent(None)
                self.table_widget.setCellWidget(row2, col, widget1)
            if widget2:
                widget2.setParent(None)
                self.table_widget.setCellWidget(row1, col, widget2)

            if widget1:
                print(f"Widget1 valid: {widget1.isWidgetType()}")
            if widget2:
                print(f"Widget2 valid: {widget2.isWidgetType()}")


    def add_rich_widget_row(self, row):
        """增加复杂的组件行，支持以下类型
        只读文本: readonly_text:
        可改文本：editable_text: 支持绑定 onchange 事件 const onchange = (row_num, col_num, row, after_change_text) => {}
        下拉框：dropdown
        复选框：checkbox:
        单选框：radio:
        全局单选框：global_radio: 纵向全局唯一
        按钮组：button_group: 支持绑定 onclick事件 const onchange = (row_num, col_num, row) => {}

        rows:
            [{
                "type": "readonly_text",
                "value": "123",
            }, {
                "type": "editable_text",
                "value": "123",
                "onchange": (row_num, col_num, row, after_change_text) => {}
            }, {
                "type": "dropdown",
                "values": ["1", "2", "3"],
                “display_values": ["1", "2", "3"]]
                "cur_index": 0
            }, {
                "type": "global_radio",
                "value": True,
            }, {
                "type": "button_group",
                "values": [{
                   "value: "测试",
                   "onclick": (row_num, col_num, row),
                }],
            }
            ]
        :return:
        """
        nex_row_index = self.table_widget.rowCount()
        self.table_widget.setRowCount(nex_row_index+1)
        for col_index, cell in enumerate(row):
            cell_type = cell.get("type")
            cell_options = cell.get("options", {})
            if cell_type == "readonly_text":
                item = QTableWidgetItem(str(cell.get("value")))
                item.setFlags(Qt.ItemIsEnabled)
                item.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
                self.table_widget.setItem(nex_row_index, col_index, item)
            elif cell_type == "editable_text":
                item = QTableWidgetItem(str(cell.get("value")))
                item.setFlags(Qt.ItemIsEditable | Qt.ItemIsEnabled)
                self.table_widget.setItem(nex_row_index, col_index, item)
            elif cell_type == "editable_int":
                params = {
                    "cur_value": int(cell.get("value")),
                    "min_num": cell_options.get("min_num"),
                    "max_num": cell_options.get("max_num"),
                    "step": cell_options.get("step"),
                    "prefix": cell_options.get("prefix"),
                    "suffix": cell_options.get("suffix"),
                    "on_change": cell_options.get("on_change"),
                }
                item = CustomSpinBox(**params)
                self.table_widget.setCellWidget(nex_row_index, col_index, item)
            elif cell_type == "editable_color":
                initial_color = QColor(cell.get("value", "#000000"))  # 默认黑色
                color_btn = ColorPickerToolButton(initial_color)
                # 绑定回调函数
                if on_change := cell_options.get("on_change"):
                    color_btn.colorChanged.connect(lambda color: on_change(color.name()))
                self.table_widget.setCellWidget(nex_row_index, col_index, color_btn)
            elif cell_type == "dropdown":
                values = cell.get("values") or cell.get("value")  # 兼容values
                font_colors = cell_options.get("font_colors") or cell_options.get("colors")
                bg_colors = cell_options.get("bg_colors")
                if cell_options.get("multi"):
                    combo_multi_box = MultiSelectComboBox(values, cur_index=cell.get("cur_index", 0), font_colors=font_colors, bg_colors=bg_colors, first_as_none=cell_options.get("first_as_none"))
                    self.table_widget.setCellWidget(nex_row_index, col_index, combo_multi_box)
                elif cell_options.get("cascader"):
                    cascader_multi_box = CascaderSelectComboBox(values, cur_index=cell.get("cur_index", [0]), first_as_none=cell_options.get("first_as_none"))
                    self.table_widget.setCellWidget(nex_row_index, col_index, cascader_multi_box)

                else:
                    combo_box = QComboBox()
                    combo_box.addItems(values)
                    combo_box.setCurrentIndex(cell.get("cur_index", 0))
                    self.table_widget.setCellWidget(nex_row_index, col_index, combo_box)
                    # 尝试字体添加颜色
                    if font_colors:
                        for i in range(len(values)):
                            color = font_colors[i]
                            if color:
                                combo_box.setItemData(i, color, role=Qt.ForegroundRole)
                    # 尝试添加背景颜色
                    if bg_colors:
                        for i in range(len(values)):
                            color = bg_colors[i]
                            if color:
                                combo_box.setItemData(i, color, role=Qt.BackgroundRole)
            elif cell_type == "checkbox":
                check_box = QCheckBox()
                check_box.setChecked(cell.get("value", False))
                # check_box.stateChanged.connect(lambda state, nex_row_index=nex_row_index,col_index=col_index: self.__global_radio_action(state, cur_row_index=nex_row_index, cur_col_index=col_index))
                self.table_widget.setCellWidget(nex_row_index, col_index, check_box)
            elif cell_type == "global_radio":
                check_box = QCheckBox()
                check_box.setChecked(cell.get("value", False))
                check_box.stateChanged.connect(lambda state, nex_row_index=nex_row_index,col_index=col_index: self.__global_radio_action(state, cur_row_index=nex_row_index, cur_col_index=col_index))
                self.table_widget.setCellWidget(nex_row_index, col_index, check_box)
            elif cell_type == "button_group":
                button_layout = QHBoxLayout()
                for button in cell.get("values"):
                    value = button.get("value")
                    onclick = button.get("onclick")
                    button_widget = QPushButton(value)
                    # button_widget.setSizePolicy(QSizePolicy.Mininum, QSizePolicy.Fixed)
                    # button_widget.setMinimumSize(100, 30)
                    this_row_values = self.get_values_by_row_index(nex_row_index)
                    button_widget.onclick = onclick
                    button_widget.clicked.connect(lambda checked, col_index=col_index, onclick=onclick, nex_row_index=nex_row_index, this_row_values=this_row_values: onclick(nex_row_index, col_index, this_row_values))
                    # button_widget.clicked.connect(lambda checked=None, onclick=onclick: onclick(nex_row_index, col_index, this_row_values))
                    button_layout.addWidget(button_widget)

                button_container = QWidget()
                button_container.setLayout(button_layout)
                self.table_widget.setRowHeight(nex_row_index, 40)
                self.table_widget.setCellWidget(nex_row_index, col_index, button_container)

    def delete_row(self, row_index):
        # 删除行
        self.table_widget.removeRow(row_index)
        # 重新绑定所有按钮的点击事件：因为删除行后，行数发生变化，导致按钮的点击事件的参数发生变化
        for i in range(self.table_widget.rowCount()):
            this_row_values = self.get_values_by_row_index(i)
            for j in range(self.table_widget.columnCount()):
                # 获取按钮
                cell_widget = self.table_widget.cellWidget(i, j)
                if cell_widget is not None:
                    buttons = cell_widget.findChildren(QPushButton)
                    if buttons:
                        for button in buttons:
                            # 获取保存的点击操作
                            if hasattr(button, "onclick"):  # 说明是按钮组里的按钮，需要重新绑定
                                # 获取保存的点击操作
                                onclick = button.onclick
                                # 重新绑定点击事件
                                button.clicked.disconnect()
                                button.clicked.connect(lambda checked, col_index=j, onclick=onclick, nex_row_index=i,
                                                              this_row_values=this_row_values: onclick(nex_row_index, col_index,
                                                                                               this_row_values))

    def __global_radio_action(self, state, **kwargs):
        cur_row_index = kwargs.get("cur_row_index")
        cur_col_index = kwargs.get("cur_col_index")
        if state == Qt.Checked:
            for row in range(self.table_widget.rowCount()):
                if row != cur_row_index and self.table_widget.cellWidget(row, cur_col_index).isChecked():
                    self.table_widget.cellWidget(row, cur_col_index).setChecked(False)

    def get_columns(self):
        labels = []
        for column in range(self.table_widget.columnCount()):
            item = self.table_widget.horizontalHeaderItem(column)
            if item is not None:
                labels.append(item.text())
        return labels

    def row_length(self):
        return self.table_widget.rowCount()

    def col_length(self):
        return self.table_widget.columnCount()

    def get_values_by_row_index(self, row_index) -> dict:
        values = {}
        for col_index, column_name in enumerate(self.get_columns()):
            values[column_name] = self.get_cell_value(row_index, col_index)
        return values

    def save_with_color(self, path, include_cols=None, exclude_cols=None):
        # 创建一个 Workbook
        wb = Workbook()
        ws = wb.active

        # 将 DataFrame 转换为 rows
        df, colors = self.get_data_as_rows_and_color()
        if include_cols:
            df = df[include_cols]
        if exclude_cols:
            df = df.drop(exclude_cols, axis=1)
        rows = dataframe_to_rows(df, index=False, header=True)

        for i, row in enumerate(rows, 1):
            ws.append(row)
            for j, cell in enumerate(ws[i], 1):
                # 获取对应的颜色
                color = None
                if i >= 2:  # 从第二行开始，即不要表头
                    color = colors[i - 2][j - 1]
                if color:
                    color = color.replace("#", "")
                    # 创建对应的 fill
                    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    # 设置单元格的背景色
                    cell.fill = fill
        # 保存 workbook
        wb.save(path)

    def save_with_color_v2(self, path, include_cols=None, exclude_cols=None):
        # 获取数据和颜色
        df, colors = self.get_data_as_rows_and_color()
        if include_cols:
            df = df[include_cols]
        if exclude_cols:
            df = df.drop(exclude_cols, axis=1)

        # 创建一个ExcelWriter对象
        writer = pd.ExcelWriter(path, engine='xlsxwriter')

        # 将DataFrame写入Excel
        df.to_excel(writer, index=False, header=True, sheet_name='Sheet1')

        # 获取xlsxwriter对象
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # 遍历颜色数组，设置单元格的颜色
        for i, row in enumerate(colors):
            for j, color in enumerate(row):
                if color:
                    color = color.replace("#", "")
                    # 创建对应的格式
                    cell_format = workbook.add_format()
                    cell_format.set_bg_color(color)
                    # 设置单元格的格式
                    worksheet.write(i + 1, j, df.iloc[i, j], cell_format)

        # 保存Excel
        writer.save()

    def save_with_color_v3(self, path, include_cols=None, exclude_cols=None, color_mapping=None):
        """
        colors: {
            QColor.name(): [1,3,4]  #  #ffffff -> [1,2,3]
        }
        """
        # 获取数据和颜色
        df = self.get_data_as_df()
        if include_cols:
            df = df[include_cols]
        if exclude_cols:
            df = df.drop(exclude_cols, axis=1)

        # 创建一个ExcelWriter对象
        writer = pd.ExcelWriter(path, engine='xlsxwriter')

        # 将DataFrame写入Excel
        df.to_excel(writer, index=False, header=True, sheet_name='Sheet1')

        # 获取xlsxwriter对象
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # 遍历颜色数组，设置单元格的颜色
        for color, cols in color_mapping.items():
            color = color.replace("#", "")
            # 创建对应的格式
            cell_format = workbook.add_format()
            cell_format.set_bg_color(color)
            # 设置列的格式
            if isinstance(cols, list):
                for col in cols:
                    worksheet.set_column(col, col, cell_format=cell_format)
            elif isinstance(cols, dict):
                for col, row_list in cols.items():
                    for row in row_list:
                        worksheet.write(row + 1, col, df.iloc[row, col], cell_format)  # +1是去掉标题行

        # 保存Excel
        writer.save()
        writer.close()

    def save_with_merged_cells(self, path):
        if self.merged_cells:
            df = self.get_data_as_df()
            with pd.ExcelWriter(path, engine='xlsxwriter') as writer:
                # 将 DataFrame 写入 Excel
                df.to_excel(writer, index=False, header=True, sheet_name='Sheet1')
                # 获取 xlsxwriter 的 workbook 和 worksheet 对象
                workbook = writer.book
                worksheet = writer.sheets['Sheet1']

                # 合并单元格（例如，合并 A1 到 B1）
                # 使用行列索引，注意索引从 0 开始
                # 读取左上角的值
                for merged_cell in self.merged_cells.iter():
                    min_row, min_col, max_row, max_col = merged_cell
                    top_left_value = df.iloc[min_row, min_col]
                    worksheet.merge_range(min_row, min_col, max_row, max_col, top_left_value)
