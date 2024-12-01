import os
import typing
import uuid
from functools import lru_cache
from multiprocessing import Pool, cpu_count, Manager

import pandas as pd
import xlrd
from openpyxl.reader.excel import load_workbook
from openpyxl.worksheet.merge import MergedCellRange


# d = Manager().dict()  # 创建一个可以在多个进程之间共享的字典


def is_empty(value):
    if pd.isna(value) or value is None:
        return True
    return len(str(value)) == 0


def is_not_empty(value):
    return not is_empty(value)


def is_any_empty(*values):
    for i in values:
        if is_empty(i):
            return True
    return False


def all_empty(*values):
    for i in values:
        if not is_empty(i):
            return False
    return True


@lru_cache(maxsize=None)
def read_excel_file(path, sheet_name, row_num_for_column, nrows, with_merged_cells, *args, **kwargs) -> pd.DataFrame:
    """
    :param path:
    :param sheet_name:
    :param row_num_for_column:
    :param nrows:
    :param with_merged_cells:
    :return: [{
        "path": "",
        "sheet_name": "",
        "row_num_for_column": 0,
        "nrows": 1,
        "with_merged_cells": False
    }]
    """
    # if str((path, sheet_name, row_num_for_column, nrows)) in d:
    #     return d[str(("df", path, sheet_name, row_num_for_column, nrows))]
    # path = file_config.get("path")
    # sheet_name = file_config.get("sheet_name")
    # header = int(file_config.get("row_num_for_column")) - 1
    # nrows = file_config.get("nrows")
    header = None
    if row_num_for_column is not None:
        header = int(row_num_for_column) - 1
    df = pd.read_excel(path, sheet_name=sheet_name, header=header, nrows=nrows)
    if with_merged_cells:
        wb = load_workbook(filename=path)
        # 选择要操作的sheet
        sheet = wb[sheet_name]
        # 获取所有合并单元格的信息
        merged_cells = sheet.merged_cells.ranges
        df.merged_cells = MergedCells(merged_cells)
    # d[str(("df", path, sheet_name, row_num_for_column, nrows))] = df
    return df


@lru_cache(maxsize=None)
def read_excel_sheets(path, *args, **kwargs) -> typing.List[str]:
    # path = file_config.get("path")
    # if str(("sheets", path, sheet_name, row_num_for_column, nrows)) in d:
    #     return d[str(("sheets", path, sheet_name, row_num_for_column, nrows))]
    sheet_names = pd.ExcelFile(path).sheet_names
    # d[str(("sheets", path, sheet_name, row_num_for_column, nrows))] = sheet_names
    return sheet_names


@lru_cache(maxsize=None)
def read_excel_columns(path, sheet_name, row_num_for_column, *args, **kwargs) -> typing.List[str]:
    # path = file_config.get("path")
    # sheet_name = file_config.get("sheet_name")
    # row_num_for_column = file_config.get("row_num_for_column")
    # if str(("columns", path, sheet_name, row_num_for_column, nrows)) in d:
    #     return d[str(("columns", path, sheet_name, row_num_for_column, nrows))]
    if path.endswith(".xlsx"):
        # 加载工作簿
        wb = load_workbook(filename=path, read_only=True)
        # 获取第一个工作表
        ws = wb[sheet_name]
        # 读取特定的行
        row_data = [ws.cell(int(row_num_for_column), col).value for col in range(1, 100) if ws.cell(1, col).value]
    else:
        # 打开工作簿
        wb = xlrd.open_workbook(path)
        # 获取第一个工作表
        sheet = wb.sheet_by_name(sheet_name)
        # 读取特定的行
        row_data = sheet.row_values(int(row_num_for_column) - 1)
    # d[str(("columns", path, sheet_name, row_num_for_column, nrows))] = row_data
    return [str(i) for i in row_data]


def read_excel_file_with_multiprocessing(file_configs, only_sheet_name=False, only_column_name=False, use_cache=True):
    """
    :param file_configs:
        [{
            "path": "",
            "sheet_name": "",
            "row_num_for_column": 0,  # 标题行
            "nrows": 1,
            "with_merged_cells": False,
        }]
    :param only_sheet_name:
    :param only_column_name:
    :param use_cache: 是否读缓存，默认是
    :return:
    """
    seed = None
    if not use_cache:
        seed = uuid.uuid4().hex
    func = read_excel_file
    if only_sheet_name:
        func = read_excel_sheets
    elif only_column_name:
        func = read_excel_columns

    if len(file_configs) == 1:
        config = file_configs[0]
        return [func(config.get("path"), config.get("sheet_name"), config.get("row_num_for_column") or 1, config.get("nrows"), config.get("with_merged_cells"), seed)]

    file_configs_list = [
        (config.get("path"), config.get("sheet_name"), config.get("row_num_for_column") or 1, config.get("nrows"), config.get("with_merged_cells"), seed) for config
        in file_configs]

    # 多于1个用多进程
    num_cores = cpu_count()  # 获取CPU的核心数
    with Pool(processes=min(num_cores, len(file_configs))) as pool:  # 创建一个包含4个进程的进程池
        results = pool.starmap(func, file_configs_list)  # 将函数和参数列表传递给进程池
    return results


MERGED_CELLS_TYPE = typing.Union[typing.List[MergedCellRange], typing.List[tuple]]


class MergedCells:
    def __init__(self, merged_cells_in_num):
        # 从1开始
        merged_cells_in_num = merged_cells_in_num or []
        self.merged_cells = merged_cells_in_num  # [(min_row, min_col, max_row, max_col), ()]
        self.inserted_cols = []  # 插入的列
        self.inserted_rows = []  # 插入的行
        self.col_merged_cell_mapping = {}  # {3: 1}  说明第三列的合并情况和第1列一样

    def insert_col(self, col_index):
        self.inserted_cols.append(col_index)
        return self

    def insert_row(self, row_index):
        self.inserted_rows.append(row_index)
        return self

    def add_col_merged_cell_mapping(self, this, copy_from):
        """添加列映射，意味着
        self.col_merged_cell_mapping[3] = 1  # 说明第3列和 第1列的合并情况一样
        """
        self.col_merged_cell_mapping[this+1] = copy_from+1

    def iter(self, index=True):
        # 修改这里的代码，使得可以反映 inserted_cols 和 inserted_rows
        # 即，如果给第0列添加了一列，那么原来跨过第0列的单元格，应该变成跨过第1列的单元格
        copy_from_col_merged_cell_mapping = {}
        for merged_cell in self.merged_cells:
            min_row, min_col, max_row, max_col = 0, 0, 0, 0
            if isinstance(merged_cell, MergedCellRange):
                min_row, min_col, max_row, max_col = merged_cell.min_row, merged_cell.min_col, merged_cell.max_row, merged_cell.max_col
            elif isinstance(merged_cell, (list, tuple)):
                min_row, min_col, max_row, max_col = merged_cell
            # 调整行索引
            for row in self.inserted_rows:
                if min_row > row+1:
                    min_row += 1
                if max_row > row+1:
                    max_row += 1

            # 调整列索引
            for col in self.inserted_cols:
                if min_col > col+1:
                    min_col += 1
                if max_col > col+1:
                    max_col += 1

            # 记录合并情况一致的列
            if min_col == max_col and min_col in self.col_merged_cell_mapping.values():
                merged_cells_this = copy_from_col_merged_cell_mapping.get(min_col, [])
                if (min_row, min_col, max_row, max_col) not in merged_cells_this:
                    merged_cells_this.append((min_row, min_col, max_row, max_col))
                copy_from_col_merged_cell_mapping[min_col] = merged_cells_this
            # 如果index是True，说明要索引，否则要数字
            yield min_row - int(index), min_col - int(index), max_row - int(index), max_col - int(index)

        if copy_from_col_merged_cell_mapping:
            for k, v in self.col_merged_cell_mapping.items():
                if v in copy_from_col_merged_cell_mapping:
                    # 说明k列需要用v列的合并
                    merged_cells_this = copy_from_col_merged_cell_mapping.get(v)
                    for merged_cell_this in merged_cells_this:
                        min_row, min_col, max_row, max_col = merged_cell_this
                        yield min_row - int(index), k - int(index), max_row - int(index), k - int(index)


def generate_unique_column_name(df, base_name, wrapper="%"):
    column_name = f"{wrapper}{base_name}{wrapper}"
    counter = 1
    while column_name in df.columns:
        column_name = f"{wrapper}{base_name}_{counter}{wrapper}"
        counter += 1
    return column_name
